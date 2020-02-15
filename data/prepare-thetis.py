# python3 ./prepare-thetis.py "2018-v127-24122019-EU MRV Publication of information.xlsx"

from openpyxl import load_workbook
import openpyxl
import csv
import sys

def get_reformatted_columns():
    with open('thetis_columns_mapping.csv', 'r') as f:
        mapping_csv = csv.DictReader(f)
        return [{'name': r['reformatted_column'], 'data_type': r['data_type']} for r in mapping_csv]


def prepare(path):
    reformatted_columns = get_reformatted_columns()
    workbook = load_workbook(filename=path)
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        with open(f"{path.split('.')[0]}.csv", 'w') as f:
            c = csv.writer(f)
            c.writerow([r['name'] for r in reformatted_columns])
            for row_idx, row in enumerate(sheet.rows):
                if row_idx < 3:
                    continue
                reformatted_row = []
                for col_idx, cell in enumerate(row):
                    if cell.value in ['Missing source values!', 'N/A', None]:
                        cell_value = None
                    elif reformatted_columns[col_idx]['data_type'] == 'float':
                        cell_value = float(cell.value)
                    else:
                        cell_value = cell.value
                    reformatted_row.append(cell_value)
                c.writerow(reformatted_row)


if __name__ == "__main__":
    print("coucou")
    prepare(sys.argv[1])
