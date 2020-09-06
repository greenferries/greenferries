# python3 scripts/convert_thetis_xlsx_to_csv.py files_original/original.thetis.export_2018.xlsx files_computed/thetis_mrv.csv

from openpyxl import load_workbook
import openpyxl
import csv
import sys
import os
DIRNAME = os.path.dirname(__file__)
MAPPING_CSV_PATH = os.path.join(
    DIRNAME,
    '../files_original/original.greenferries.thetis_columns_mapping.csv'
)

class Converter():

    def __init__(self, xlsx_path, csv_path):
        self.xlsx_path = xlsx_path
        self.csv_path = csv_path
        self.reformatted_columns = None
        pass

    def get_reformatted_columns(self):
        if self.reformatted_columns is None:
            with open(MAPPING_CSV_PATH, 'r') as f:
                mapping_csv = csv.DictReader(f)
                self.reformatted_columns = [
                    {'name': r['reformatted_column'], 'data_type': r['data_type']}
                    for r in mapping_csv
                ]
        return self.reformatted_columns

    def run(self):
        workbook = load_workbook(filename=self.xlsx_path)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            with open(self.csv_path, 'w') as f:
                c = csv.writer(f)
                c.writerow([r['name'] for r in self.get_reformatted_columns()])
                for row_idx, row in enumerate(sheet.rows):
                    if row_idx < 3:
                        continue
                    reformatted_row = []
                    for col_idx, cell in enumerate(row):
                        if cell.value in ['Missing source values!', 'N/A', None]:
                            cell_value = None
                        elif self.get_reformatted_columns()[col_idx]['data_type'] == 'float':
                            cell_value = float(cell.value)
                        else:
                            cell_value = cell.value
                        reformatted_row.append(cell_value)
                    c.writerow(reformatted_row)


if __name__ == "__main__":
    Converter(sys.argv[1], sys.argv[2]).run()
