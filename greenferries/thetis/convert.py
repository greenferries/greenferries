# python3 -m greenferries.thetis.convert

from openpyxl import load_workbook
import openpyxl, csv, sys, os, re
from greenferries.utils import run_sh, clean_file

PACKAGE_ROOT = os.path.join(os.path.dirname(__file__), "..")
MAPPING_CSV_PATH = os.path.join(
    PACKAGE_ROOT,
    'files_original/original.greenferries.thetis_columns_mapping.csv'
)


class Convert():
    def __init__(self):
        self.reformatted_columns = None
        pass

    def get_reformatted_columns(self):
        if self.reformatted_columns is None:
            with open(MAPPING_CSV_PATH, 'r') as f:
                mapping_csv = csv.DictReader(f)
                self.reformatted_columns = [
                    {'name': r['reformatted_column'], 'data_type': r['data_type']}
                    for r in mapping_csv
                ]
        return self.reformatted_columns

    def run(self):
        for filename in os.listdir(os.path.join(PACKAGE_ROOT, "files_original")):
            match_data = re.match(r"original\.thetis\.export_(\d+)\.xlsx", filename)
            if not match_data:
                continue
            year = match_data.groups()[0]
            csv_filename = "thetis_export_%s.csv" % year
            xlsx_path = os.path.join(PACKAGE_ROOT, "files_original", filename)
            csv_path = os.path.join(PACKAGE_ROOT, "files_computed", csv_filename)
            self.convert(xlsx_path, csv_path)
        self.join_csvs()

    def convert(self, xlsx_path, csv_path):
        workbook = load_workbook(filename=xlsx_path)
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            with open(csv_path, 'w') as f:
                c = csv.writer(f)
                c.writerow([r['name'] for r in self.get_reformatted_columns()])
                for row_idx, row in enumerate(sheet.rows):
                    if row_idx < 3:
                        continue
                    reformatted_row = []
                    for col_idx, cell in enumerate(row):
                        if cell.value in ['Missing source values!', 'N/A', 'Division by zero!', None]:
                            cell_value = None
                        elif self.get_reformatted_columns()[col_idx]['data_type'] == 'float':
                            cell_value = float(cell.value)
                        else:
                            cell_value = cell.value
                        reformatted_row.append(cell_value)
                    c.writerow(reformatted_row)
            print("rewrote %s" % csv_path)

    def join_csvs(self):
        glob_path = os.path.join(PACKAGE_ROOT, "files_computed", "thetis_export_*.csv")
        joined_path = os.path.join(PACKAGE_ROOT, "files_computed", "thetis_export_all.csv")
        clean_file(joined_path)
        run_sh(f"awk '(NR == 1) || (FNR > 1)' {glob_path} > {joined_path}")

if __name__ == "__main__":
    Convert().run()
